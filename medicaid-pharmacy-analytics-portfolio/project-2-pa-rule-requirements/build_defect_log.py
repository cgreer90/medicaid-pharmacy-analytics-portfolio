import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/claude/ba_project/PA_GLP1_Test_Case_Matrix.xlsx')
ws = wb['Test Case Matrix']

DARK_BLUE = "1F3864"
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb_vals = openpyxl.load_workbook('/home/claude/ba_project/PA_GLP1_Test_Case_Matrix.xlsx', data_only=True)
ws_vals = wb_vals['Test Case Matrix']

fail_rows = []
for r in range(2, 27):
    result = ws_vals.cell(row=r, column=12).value
    cell = ws.cell(row=r, column=12)
    if result == "FAIL":
        cell.fill = FAIL_FILL
        fail_rows.append(r)
    elif result == "PASS":
        cell.fill = PASS_FILL

# ---------------- Defect Log sheet ----------------
dl = wb.create_sheet("Defect Log")
headers = ["Defect ID", "Related Test Case", "Description", "Expected Behavior", "Actual (Observed) Behavior",
           "Severity", "Root Cause (Hypothesis)", "Recommended Resolution", "Status"]
for i, h in enumerate(headers, start=1):
    dl.cell(row=1, column=i, value=h)
    dl.cell(row=1, column=i).fill = HEADER_FILL
    dl.cell(row=1, column=i).font = HEADER_FONT
    dl.cell(row=1, column=i).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dl.cell(row=1, column=i).border = BORDER
dl.freeze_panes = "A2"

defects = [
    ("DEF-001", "TC-05", "Claim with quantity billed (2) exceeding the 1-fill/28-day limit was approved instead of denied.",
     "Deny — quantity exceeds the defined limit.", "Approve",
     "High", "Quantity-limit edit does not appear to be evaluated for this drug category, or is being bypassed when other criteria (diagnosis, BMI, step therapy) are met.",
     "Verify quantity-limit edit configuration for GLP-1 weight-management NDCs specifically; confirm the edit fires independently of other approval criteria rather than being short-circuited by them.",
     "Open"),
    ("DEF-002", "TC-11", "Claim with a missing/blank BMI value on the PA record was denied instead of pended for additional information.",
     "Pend — required field (BMI) is missing.", "Deny",
     "Medium", "System may be treating a blank/null BMI field as a failed numeric comparison (e.g., blank < 30 evaluates as true) rather than first checking for missing data.",
     "Add an explicit null/blank check for BMI (and other required fields) ahead of the numeric threshold comparison so missing data routes to Pend before any Approve/Deny logic runs.",
     "Open"),
    ("DEF-003", "TC-14", "Claim submitted with a Type 2 diabetes diagnosis code was approved under the weight-management rule instead of being denied as out of scope.",
     "Deny — diagnosis code is not on the approved weight-management list.", "Approve",
     "High", "The approved-diagnosis list check may not be restricting correctly, or the diabetes-indication claim is being evaluated against the wrong rule set entirely (routing issue rather than a logic issue within this rule).",
     "Confirm claim routing correctly separates diabetes-indication GLP-1 claims (existing diabetes PA rule) from weight-management-indication claims (this rule) before this rule's logic is applied; add a regression test covering diagnosis-code routing.",
     "Open"),
]

r = 2
for d in defects:
    for i, val in enumerate(d, start=1):
        dl.cell(row=r, column=i, value=val)
        dl.cell(row=r, column=i).font = BODY_FONT
        dl.cell(row=r, column=i).border = BORDER
        dl.cell(row=r, column=i).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

col_widths = [10, 14, 34, 26, 20, 10, 34, 34, 10]
for i, w in enumerate(col_widths, start=1):
    dl.column_dimensions[get_column_letter(i)].width = w
for rr in range(2, r):
    dl.row_dimensions[rr].height = 60

wb.save('/home/claude/ba_project/PA_GLP1_Test_Case_Matrix.xlsx')
print("Defect Log sheet added with", len(defects), "defects. Fail rows:", fail_rows)
