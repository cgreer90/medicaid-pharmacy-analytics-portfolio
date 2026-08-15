import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Case Matrix"

DARK_BLUE = "1F3864"
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    "Test Case ID", "Scenario Description", "Diagnosis Code", "Approved Dx?",
    "BMI", "Age", "Step Therapy on File?", "Quantity Billed", "Days Supply",
    "System Expected Outcome", "Actual System Outcome (Observed)", "Result"
]
for i, h in enumerate(headers, start=1):
    ws.cell(row=1, column=i, value=h)
    ws.cell(row=1, column=i).fill = HEADER_FILL
    ws.cell(row=1, column=i).font = HEADER_FONT
    ws.cell(row=1, column=i).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=1, column=i).border = BORDER
ws.freeze_panes = "A2"

APPROVED_DX = {"E66.01", "E66.09", "E66.3"}

# Test cases: (id, description, dx, bmi, age, step_therapy, qty, days_supply, actual_observed)
# bmi/age/step_therapy/dx can be None to represent missing data (triggers PEND)
cases = [
    ("TC-01", "Straightforward approve — meets all criteria", "E66.01", 32, 34, "Y", 1, 28, "Approve"),
    ("TC-02", "Straightforward deny — BMI below threshold", "E66.01", 27, 40, "Y", 1, 28, "Deny"),
    ("TC-03", "Deny — diagnosis not on approved list (unrelated dx)", "R63.5", 33, 29, "Y", 1, 28, "Deny"),
    ("TC-04", "Deny — step therapy not on file", "E66.09", 31, 45, "N", 1, 28, "Deny"),
    ("TC-05", "Deny — quantity exceeds limit", "E66.3", 35, 38, "Y", 2, 28, "Approve"),
    ("TC-06", "Deny — under age 18", "E66.01", 31, 16, "Y", 1, 28, "Deny"),
    ("TC-07", "Edge case — BMI exactly at threshold (30)", "E66.09", 30, 41, "Y", 1, 28, "Approve"),
    ("TC-08", "Edge case — age exactly at threshold (18)", "E66.3", 31, 18, "Y", 1, 28, "Approve"),
    ("TC-09", "Edge case — quantity exactly at limit (1)", "E66.01", 34, 50, "Y", 1, 28, "Approve"),
    ("TC-10", "Pend — missing diagnosis code", None, 32, 33, "Y", 1, 28, "Pend"),
    ("TC-11", "Pend — missing BMI on PA record", "E66.09", None, 29, "Y", 1, 28, "Deny"),
    ("TC-12", "Pend — missing age on record", "E66.01", 33, None, "Y", 1, 28, "Pend"),
    ("TC-13", "Pend — missing step-therapy indicator", "E66.3", 36, 44, None, 1, 28, "Pend"),
    ("TC-14", "Diabetes diagnosis submitted — out of scope for this rule, should deny under weight-mgmt rule", "E11.9", 33, 47, "Y", 1, 28, "Approve"),
    ("TC-15", "Deny — days supply mismatch with 28-day fill logic", "E66.01", 32, 39, "Y", 1, 45, "Deny"),
    ("TC-16", "Approve — high BMI, all criteria clearly met", "E66.3", 41, 52, "Y", 1, 28, "Approve"),
    ("TC-17", "Deny — quantity well above limit (bulk fill)", "E66.09", 33, 36, "Y", 3, 28, "Deny"),
    ("TC-18", "Approve — minimum valid days supply", "E66.01", 31, 30, "Y", 1, 28, "Approve"),
    ("TC-19", "Deny — BMI just below threshold (29)", "E66.09", 29, 42, "Y", 1, 28, "Deny"),
    ("TC-20", "Approve — older adult member", "E66.3", 33, 64, "Y", 1, 28, "Approve"),
    ("TC-21", "Pend — multiple fields missing (dx and BMI both blank)", None, None, 37, "Y", 1, 28, "Pend"),
    ("TC-22", "Deny — step therapy not on file, otherwise qualifies", "E66.3", 38, 55, "N", 1, 28, "Deny"),
    ("TC-23", "Approve — second qualifying diagnosis code variant", "E66.09", 30, 33, "Y", 1, 28, "Approve"),
    ("TC-24", "Deny — age below 18 with otherwise valid record", "E66.3", 34, 17, "Y", 1, 28, "Deny"),
    ("TC-25", "Approve — days supply at upper valid bound (31)", "E66.01", 32, 46, "Y", 1, 31, "Approve"),
]

r = 2
for c in cases:
    tc_id, desc, dx, bmi, age, step, qty, days, actual = c
    ws.cell(row=r, column=1, value=tc_id)
    ws.cell(row=r, column=2, value=desc)
    ws.cell(row=r, column=3, value=dx if dx is not None else "")
    ws.cell(row=r, column=4, value=f'=IF(C{r}="","N/A",IF(OR(C{r}="E66.01",C{r}="E66.09",C{r}="E66.3"),"Y","N"))')
    ws.cell(row=r, column=5, value=bmi if bmi is not None else "")
    ws.cell(row=r, column=6, value=age if age is not None else "")
    ws.cell(row=r, column=7, value=step if step is not None else "")
    ws.cell(row=r, column=8, value=qty)
    ws.cell(row=r, column=9, value=days)
    # System Expected Outcome formula per BRD logic
    formula = (
        f'=IF(OR(C{r}="",E{r}="",F{r}="",G{r}=""),"Pend",'
        f'IF(AND(D{r}="Y",E{r}>=30,F{r}>=18,G{r}="Y",H{r}<=1,I{r}>=28,I{r}<=31),"Approve","Deny"))'
    )
    ws.cell(row=r, column=10, value=formula)
    ws.cell(row=r, column=11, value=actual)
    ws.cell(row=r, column=12, value=f'=IF(J{r}=K{r},"PASS","FAIL")')
    for c_idx in range(1, 13):
        ws.cell(row=r, column=c_idx).font = BODY_FONT
        ws.cell(row=r, column=c_idx).border = BORDER
    r += 1

last_row = r - 1

# Color-code the Result column based on the actual pre-set expectation (visual aid; real color applied post-recalc separately)
# Conditional-style fill for Result column based on value (applied after recalculation in a pass)
col_widths = [10, 42, 12, 10, 6, 6, 12, 10, 10, 16, 18, 9]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.cell(row=last_row + 2, column=1, value="Legend: PASS = System Expected Outcome matches Actual System Outcome (Observed). FAIL = mismatch — logged as a defect (see Defect Log tab).")
ws.cell(row=last_row + 2, column=1).font = Font(italic=True, size=9, color="666666", name="Arial")
ws.merge_cells(start_row=last_row + 2, start_column=1, end_row=last_row + 2, end_column=12)

wb.save("/home/claude/ba_project/PA_GLP1_Test_Case_Matrix.xlsx")
print("Test case matrix built:", last_row - 1, "test cases")
