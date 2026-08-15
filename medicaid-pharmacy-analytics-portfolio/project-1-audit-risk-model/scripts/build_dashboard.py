import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/home/claude/mi_audit/MI_Medicaid_Audit_Workbook.xlsx')

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=11, bold=True)
TITLE_FONT = Font(name="Arial", size=16, bold=True, color="1F3864")
FLAG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

RS_LAST = 201  # Risk Scoring data rows 2..201

# ---------------- High-Risk Audit Sample (Top 30) ----------------
ws = wb.create_sheet("High-Risk Audit Sample")
headers = ["Audit Priority Rank", "Product Name", "Total Prescriptions", "Total Amount Reimbursed",
           "Avg Cost per Rx", "Utilization Volatility", "Risk Score", "Primary Risk Driver"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))
ws.freeze_panes = "A2"

TOP_N = 30
for k in range(1, TOP_N + 1):
    r = k + 1
    ws.cell(row=r, column=1, value=k)
    ws.cell(row=r, column=2, value=f"=INDEX('Risk Scoring'!$A$2:$A${RS_LAST},MATCH($A{r},'Risk Scoring'!$O$2:$O${RS_LAST},0))")
    ws.cell(row=r, column=3, value=f"=INDEX('Risk Scoring'!$B$2:$B${RS_LAST},MATCH($A{r},'Risk Scoring'!$O$2:$O${RS_LAST},0))")
    ws.cell(row=r, column=4, value=f"=INDEX('Risk Scoring'!$D$2:$D${RS_LAST},MATCH($A{r},'Risk Scoring'!$O$2:$O${RS_LAST},0))")
    ws.cell(row=r, column=5, value=f"=INDEX('Risk Scoring'!$E$2:$E${RS_LAST},MATCH($A{r},'Risk Scoring'!$O$2:$O${RS_LAST},0))")
    ws.cell(row=r, column=6, value=f"=INDEX('Risk Scoring'!$K$2:$K${RS_LAST},MATCH($A{r},'Risk Scoring'!$O$2:$O${RS_LAST},0))")
    ws.cell(row=r, column=7, value=f"=INDEX('Risk Scoring'!$N$2:$N${RS_LAST},MATCH($A{r},'Risk Scoring'!$O$2:$O${RS_LAST},0))")
    ws.cell(row=r, column=8, value=(
        f'=IF(ABS(INDEX(\'Risk Scoring\'!$L$2:$L${RS_LAST},MATCH($A{r},\'Risk Scoring\'!$O$2:$O${RS_LAST},0)))'
        f'>=ABS(INDEX(\'Risk Scoring\'!$M$2:$M${RS_LAST},MATCH($A{r},\'Risk Scoring\'!$O$2:$O${RS_LAST},0))),'
        f'"Cost Outlier","Utilization Volatility")'
    ))
    for c in range(1, 9):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).border = BORDER
        if k <= 10:
            ws.cell(row=r, column=c).fill = FLAG_FILL
    ws.cell(row=r, column=3).number_format = '#,##0'
    ws.cell(row=r, column=4).number_format = '$#,##0.00'
    ws.cell(row=r, column=5).number_format = '$#,##0.00'
    ws.cell(row=r, column=6).number_format = '0.00%'
    ws.cell(row=r, column=7).number_format = '0.00'

col_widths = [8, 18, 16, 18, 14, 16, 11, 18]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

note_row = TOP_N + 3
ws.cell(row=note_row, column=1, value=(
    "Methodology: Risk Score = Cost Z-Score + Utilization Volatility Z-Score, computed across the top 200 "
    "products by prescription volume in Michigan Medicaid, 2024 (all quarters). Top 10 rows highlighted "
    "represent the highest-priority audit candidates. See Methodology tab for full detail."
))
ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, name="Arial", color="666666")
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)

# ---------------- Dashboard ----------------
dash = wb.create_sheet("Dashboard", 0)
dash.sheet_view.showGridLines = False
dash.column_dimensions['A'].width = 3
for col, w in zip("BCDEFGH", [26, 18, 18, 18, 18, 18, 18]):
    dash.column_dimensions[col].width = w

dash.cell(row=2, column=2, value="Michigan Medicaid Pharmacy Claims — Audit Risk Dashboard")
dash.cell(row=2, column=2).font = TITLE_FONT
dash.merge_cells("B2:H2")

dash.cell(row=3, column=2, value="Data Source: CMS State Drug Utilization Data (SDUD), Michigan, FY2024, Quarters 1–4")
dash.cell(row=3, column=2).font = Font(italic=True, size=10, name="Arial", color="666666")
dash.merge_cells("B3:H3")

kpi_labels = [
    ("Total Claim-Lines Analyzed", "=ROWS('Raw Data'!$A$2:$A$131358)"),
    ("Total Reimbursed (All MI Rx, FY24)", "=SUM('Raw Data'!$M$2:$M$131358)"),
    ("Products in Audit Universe (Top by Volume)", "=COUNTA('Risk Scoring'!$A$2:$A$201)"),
    ("High-Risk Products Flagged for Audit", "=COUNTA('High-Risk Audit Sample'!$B$2:$B$31)"),
    ("Total $ Represented by Flagged Sample", "=SUM('High-Risk Audit Sample'!$D$2:$D$31)"),
    ("Flagged Sample as % of Audit Universe $", "=IFERROR(F7/SUM('Risk Scoring'!$D$2:$D$201),0)"),
]

row = 5
for label, formula in kpi_labels:
    dash.cell(row=row, column=2, value=label).font = BOLD_FONT
    cell = dash.cell(row=row, column=5, value=formula)
    cell.font = Font(name="Arial", size=12, bold=True, color="1F3864")
    if "Total Reimbursed" in label or "Total $" in label:
        cell.number_format = '$#,##0'
    elif "%" in label:
        cell.number_format = '0.0%'
    else:
        cell.number_format = '#,##0'
    row += 1

# Bar chart: Top 15 risk scores
chart = BarChart()
chart.title = "Top 15 Highest-Risk Products (by Risk Score)"
chart.style = 10
chart.y_axis.title = 'Risk Score'
chart.x_axis.title = 'Product'
data = Reference(wb['High-Risk Audit Sample'], min_col=7, min_row=1, max_row=16)
cats = Reference(wb['High-Risk Audit Sample'], min_col=2, min_row=2, max_row=16)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 24
chart.height = 12
dash.add_chart(chart, "B13")

wb.save('/home/claude/mi_audit/MI_Medicaid_Audit_Workbook.xlsx')
print("High-Risk Audit Sample and Dashboard sheets built.")
