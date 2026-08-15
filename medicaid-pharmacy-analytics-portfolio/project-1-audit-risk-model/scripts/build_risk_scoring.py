import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import pandas as pd

wb = openpyxl.load_workbook('/home/claude/mi_audit/MI_Medicaid_Audit_Workbook.xlsx')
top200 = pd.read_csv('/home/claude/top200_products.csv')

RAW_LAST_ROW = 131358  # header=1, data rows 2..131358

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
FLAG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

# Style Raw Data header
ws_raw = wb['Raw Data']
style_header(ws_raw, 1, 15)
ws_raw.freeze_panes = "A2"
col_widths_raw = [16, 6, 14, 12, 12, 12, 7, 8, 12, 14, 14, 16, 16, 16, 18]
for i, w in enumerate(col_widths_raw, start=1):
    ws_raw.column_dimensions[get_column_letter(i)].width = w

# ---------------- Risk Scoring sheet ----------------
ws = wb.create_sheet("Risk Scoring")
headers = [
    "Product Name", "Total Prescriptions", "Total Units Reimbursed", "Total Amount Reimbursed",
    "Avg Cost per Rx", "Avg Cost per Unit", "Q1 Rx", "Q2 Rx", "Q3 Rx", "Q4 Rx",
    "Utilization Volatility", "Cost Z-Score", "Volatility Z-Score", "Risk Score", "Risk Rank"
]
for i, h in enumerate(headers, start=1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))
ws.freeze_panes = "A2"

n = len(top200)
last_data_row = 1 + n  # row 2..201

for idx, row in enumerate(top200.itertuples(), start=2):
    ws.cell(row=idx, column=1, value=row._1)  # Product Name
    # Total Prescriptions
    ws.cell(row=idx, column=2, value=f"=SUMIFS('Raw Data'!$L$2:$L${RAW_LAST_ROW},'Raw Data'!$J$2:$J${RAW_LAST_ROW},$A{idx})")
    # Total Units Reimbursed
    ws.cell(row=idx, column=3, value=f"=SUMIFS('Raw Data'!$K$2:$K${RAW_LAST_ROW},'Raw Data'!$J$2:$J${RAW_LAST_ROW},$A{idx})")
    # Total Amount Reimbursed
    ws.cell(row=idx, column=4, value=f"=SUMIFS('Raw Data'!$M$2:$M${RAW_LAST_ROW},'Raw Data'!$J$2:$J${RAW_LAST_ROW},$A{idx})")
    # Avg Cost per Rx
    ws.cell(row=idx, column=5, value=f"=IFERROR($D{idx}/$B{idx},0)")
    # Avg Cost per Unit
    ws.cell(row=idx, column=6, value=f"=IFERROR($D{idx}/$C{idx},0)")
    # Q1-Q4 Rx
    for qi, qcol in enumerate([7, 8, 9, 10], start=1):
        ws.cell(row=idx, column=qcol, value=f"=SUMIFS('Raw Data'!$L$2:$L${RAW_LAST_ROW},'Raw Data'!$J$2:$J${RAW_LAST_ROW},$A{idx},'Raw Data'!$H$2:$H${RAW_LAST_ROW},{qi})")
    # Utilization Volatility = (max-min)/avg across quarters
    ws.cell(row=idx, column=11, value=f"=IFERROR((MAX(G{idx}:J{idx})-MIN(G{idx}:J{idx}))/AVERAGE(G{idx}:J{idx}),0)")
    # Cost Z-Score (population z-score across the 200-product cohort)
    ws.cell(row=idx, column=12, value=f"=IFERROR((E{idx}-AVERAGE($E$2:$E${last_data_row}))/STDEV($E$2:$E${last_data_row}),0)")
    # Volatility Z-Score
    ws.cell(row=idx, column=13, value=f"=IFERROR((K{idx}-AVERAGE($K$2:$K${last_data_row}))/STDEV($K$2:$K${last_data_row}),0)")
    # Risk Score = sum of the two z-scores
    ws.cell(row=idx, column=14, value=f"=L{idx}+M{idx}")
    # Risk Rank (1 = highest risk)
    ws.cell(row=idx, column=15, value=f"=RANK(N{idx},$N$2:$N${last_data_row},0)")
    for c in range(1, 16):
        ws.cell(row=idx, column=c).font = BODY_FONT
        ws.cell(row=idx, column=c).border = BORDER

money_fmt = '$#,##0.00'
num_fmt = '#,##0'
for r in range(2, last_data_row + 1):
    for c in [3]:
        ws.cell(row=r, column=c).number_format = num_fmt
    for c in [4, 5, 6]:
        ws.cell(row=r, column=c).number_format = money_fmt
    for c in [2, 7, 8, 9, 10]:
        ws.cell(row=r, column=c).number_format = num_fmt
    ws.cell(row=r, column=11).number_format = '0.00%'
    for c in [12, 13, 14]:
        ws.cell(row=r, column=c).number_format = '0.00'

col_widths = [16, 16, 18, 18, 14, 14, 9, 9, 9, 9, 16, 12, 14, 11, 10]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save('/home/claude/mi_audit/MI_Medicaid_Audit_Workbook.xlsx')
print("Risk Scoring sheet built:", n, "products")
